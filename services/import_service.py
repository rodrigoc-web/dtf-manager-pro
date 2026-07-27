"""
services/import_service.py — Importação local de pedidos (Excel/CSV).
Substitui a leitura do Google Sheets: o usuário mantém a planilha no
próprio computador e importa quando quiser.

O layout esperado depende do tipo do modelo casado pela coluna
profissão/time — Profissão precisa de "Telefone"; Time precisa de
"Nome", "Número Peito" e "Número Costas". Um arquivo pode até misturar
os dois (cada linha só exige as colunas do seu próprio tipo).

Cabeçalhos aceitos (sem diferenciar maiúsculas/acentos):
  telefone       → "telefone", "fone", "celular", "whatsapp", "phone"
  profissao      → "profissao", "profissão", "modelo", "categoria", "time"
  nome           → "nome", "jogador", "nome jogador"
  numero_peito   → "numero peito", "peito", "numero frente", "frente"
  numero_costas  → "numero costas", "costas", "numero"
  operador       → "operador", "usuario", "responsavel", "produzido por"
  quantidade     → "quantidade", "qtd", "qtde"  (opcional, padrão 1)
"""
from __future__ import annotations
import csv
import unicodedata
from pathlib import Path
from domain.models import Pedido, Modelo
from core.constants import CAMPOS_POR_TIPO
from core.exceptions import ImportacaoError

_ALIASES = {
    "telefone":      ["telefone", "fone", "celular", "whatsapp", "phone"],
    "profissao":     ["profissao", "modelo", "categoria", "arte", "time"],
    "nome":          ["nome", "jogador", "nome jogador"],
    "numero_peito":  ["numero peito", "num peito", "peito", "numero frente", "frente"],
    "numero_costas": ["numero costas", "num costas", "costas", "numero"],
    "operador":      ["operador", "usuario", "responsavel", "produzido por", "quem produz"],
    "marketplace":   ["marketplace", "origem", "canal", "canal de venda", "loja"],
    "quantidade":    ["quantidade", "qtd", "qtde", "quant"],
}


def _normalizar(texto: str) -> str:
    texto = unicodedata.normalize("NFKD", str(texto)).encode("ascii", "ignore").decode()
    return texto.strip().lower()


def importar_arquivo(path: str) -> list[dict]:
    """Lê CSV ou XLSX e retorna uma lista de dicts (chave = cabeçalho original)."""
    ext = Path(path).suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return _importar_xlsx(path)
    if ext == ".csv":
        return _importar_csv(path)
    raise ImportacaoError(f"Formato não suportado: {ext} (use .csv, .xlsx ou .xlsm)")


def _importar_csv(path: str) -> list[dict]:
    try:
        with open(path, encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f))
    except Exception as e:
        raise ImportacaoError(f"Erro ao ler CSV: {e}")


def _importar_xlsx(path: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        linhas = ws.iter_rows(values_only=True)
        cabecalho = [str(h).strip() if h is not None else "" for h in next(linhas)]
        resultado = []
        for linha in linhas:
            if linha is None or all(v is None for v in linha):
                continue
            resultado.append({
                cabecalho[i]: linha[i]
                for i in range(len(cabecalho)) if i < len(linha)
            })
        return resultado
    except ImportacaoError:
        raise
    except Exception as e:
        raise ImportacaoError(f"Erro ao ler Excel: {e}")


def detectar_mapeamento(cabecalhos) -> dict[str, str]:
    """Casa os cabeçalhos do arquivo com os campos esperados por heurística de nome."""
    normalizados = {_normalizar(h): h for h in cabecalhos}
    mapeamento: dict[str, str] = {}
    for campo, aliases in _ALIASES.items():
        for alias in aliases:
            if alias in normalizados:
                mapeamento[campo] = normalizados[alias]
                break
    return mapeamento


def converter_para_pedidos(
        linhas: list[dict], mapeamento: dict[str, str],
        modelos: list[Modelo]) -> tuple[list[Pedido], list[str]]:
    """
    Converte linhas brutas em Pedido, casando a profissão/time pelo nome
    cadastrado. Retorna (pedidos_validos, mensagens_de_erro) — nunca
    levanta exceção por linha individual, para não travar a importação
    inteira por causa de uma linha ruim.
    """
    if "profissao" not in mapeamento:
        raise ImportacaoError(
            "Não encontrei a coluna 'profissão'/'modelo'/'time' no arquivo.\n"
            "Renomeie o cabeçalho e tente novamente.")
    if "telefone" not in mapeamento and not all(
            k in mapeamento for k in ("nome", "numero_peito", "numero_costas")):
        raise ImportacaoError(
            "Não encontrei colunas suficientes.\n"
            "Para Profissão: coluna 'Telefone'.\n"
            "Para Time: colunas 'Nome', 'Número Peito' e 'Número Costas'.")

    por_nome = {_normalizar(m.profissao): m for m in modelos}
    pedidos: list[Pedido] = []
    erros: list[str] = []

    for i, linha in enumerate(linhas, start=2):  # linha 1 = cabeçalho
        profissao = str(linha.get(mapeamento["profissao"], "") or "").strip()
        if not profissao:
            erros.append(f"Linha {i}: profissão/time em branco — ignorada.")
            continue

        modelo = por_nome.get(_normalizar(profissao))
        if not modelo:
            erros.append(
                f"Linha {i}: '{profissao}' não está cadastrado em Modelos — ignorada.")
            continue

        campos_necessarios = CAMPOS_POR_TIPO[modelo.tipo.value]
        dados = {chave: str(linha.get(mapeamento[chave], "") or "").strip()
                 for chave, _ in campos_necessarios if chave in mapeamento}
        faltando = [rotulo for chave, rotulo in campos_necessarios if not dados.get(chave)]
        if faltando:
            erros.append(
                f"Linha {i}: faltam as colunas {', '.join(faltando)} para "
                f"'{modelo.profissao}' ({modelo.tipo.value}) — ignorada.")
            continue

        quantidade = 1
        if "quantidade" in mapeamento:
            bruto = linha.get(mapeamento["quantidade"])
            try:
                quantidade = max(1, int(float(bruto)))
            except (TypeError, ValueError):
                quantidade = 1

        operador = ""
        if "operador" in mapeamento:
            operador = str(linha.get(mapeamento["operador"], "") or "").strip()

        marketplace = ""
        if "marketplace" in mapeamento:
            marketplace = str(linha.get(mapeamento["marketplace"], "") or "").strip()

        pedidos.append(Pedido(
            id=None, modelo_id=modelo.id, profissao=modelo.profissao,
            dados=dados, operador=operador, marketplace=marketplace, quantidade=quantidade))

    return pedidos, erros
